"""
ingestion/scrapers/sbi_mf_holdings.py

Phase: 2.2 (AMFI MF Holdings + Corporate Action Features)
Specs: SPEC-PIPE-001, SPEC-PIPE-003 (CRITICAL), SPEC-MFHOLD-001
Owner: Platform / Ingestion
Consumers: ingestion/scrapers/amfi_holdings.py (registry), features/mf_holdings.py

SBI Mutual Fund's direct portfolio-disclosure scraper — registered as a
secondary, higher-precision cross-check source (real ISIN + real share
quantity) alongside Groww (`groww_mf_holdings.py`, the primary source
covering all 49 AMCs — see SPEC-MFHOLD-001 in
alphalens_docs/specs/08_specifications.md for the full sourcing decision).

Verified live: `https://www.sbimf.com/portfolios` is a JS-driven filter
form (Category / Frequency / Year / Month, all custom-styled `<select>`s
— see `ingestion/scrapers/browser.py`'s `set_select_by_label`) that, once
filled, reveals direct `.xlsx` download links embedded as real `<a href>`
elements (no further JS needed to fetch the file itself — a plain
`requests.get` on the discovered URL works). One workbook covers ALL of
SBI MF's schemes for that month: an "Index" sheet (scheme code -> short
code -> full name) plus one sheet per scheme, each with a
"SCHEME NAME :" header row and a holdings table (Name of the Instrument /
Issuer, ISIN, Rating/Industry, Quantity, Market value (Rs. in Lakhs),
% to AUM, ...) starting a few rows below. Section header rows (e.g.
"EQUITY & EQUITY RELATED") have no ISIN — filtered out by requiring a
real ISIN that resolves to a known universe ticker.

Unlike Groww, SBI's own archive supports arbitrary historical months
(real year/month dropdown selection) — the natural fallback if a past
month is ever needed for SBI specifically.

PIT Assumptions
----------------
None beyond what ingestion/scrapers/amfi_holdings.py's
save_monthly_parquet already enforces (availability_date stamping). This
module returns whatever data SBI's archive has for the exact requested
(year, month) — no PIT filtering of its own is needed since the archive
is keyed by the real disclosure period already.
"""

import calendar
import io
import logging

import openpyxl
import pandas as pd
import requests

from config.settings import AMFI_RAW_DIR
from config.universe import get_isin_to_ticker_map
from ingestion.scrapers.amfi_holdings import register_amc
from ingestion.scrapers.browser import DEFAULT_USER_AGENT, browser_page, set_select_by_label

logger = logging.getLogger(__name__)

PORTFOLIO_URL = "https://www.sbimf.com/portfolios"
AMC_NAME = "SBI Mutual Fund (Direct, ISIN-exact)"


def fetch(year: int, month: int) -> bytes:
    """
    Discover (via Playwright) and download SBI Mutual Fund's monthly
    all-schemes Excel workbook for the requested (year, month).

    Parameters
    ----------
    year : int
    month : int

    Returns
    -------
    bytes
        The raw .xlsx file content.

    Spec References
    ----------------
    SPEC-PIPE-001: raw retention.

    Raises
    ------
    ConnectionError
        If no matching download link is found for the requested month.
    requests.HTTPError
        If the file download itself fails.
    """
    month_name = calendar.month_name[month]
    with browser_page() as page:
        page.goto(PORTFOLIO_URL, timeout=30000, wait_until="domcontentloaded")
        page.wait_for_timeout(2000)
        set_select_by_label(page, "#FSCategeory", "Equity")
        page.wait_for_timeout(300)
        set_select_by_label(page, "#PSFrequency", "Monthly")
        page.wait_for_timeout(300)
        set_select_by_label(page, "#PSYear", str(year))
        page.wait_for_timeout(300)
        set_select_by_label(page, "#PSMonth", month_name)
        page.wait_for_timeout(2500)
        links = page.eval_on_selector_all(
            "a", r"els => els.map(e => e.href).filter(h => h.match(/\.xlsx(\?|$)/i))"
        )

    matches = [link for link in links if str(year) in link and month_name.lower() in link.lower()]
    if not matches:
        raise ConnectionError(f"SBI Mutual Fund: no portfolio file link found for {month_name} {year}")

    response = requests.get(matches[0], timeout=60, headers={"User-Agent": DEFAULT_USER_AGENT})
    response.raise_for_status()

    raw_dir = AMFI_RAW_DIR / "sbi_mutual_fund"
    raw_dir.mkdir(parents=True, exist_ok=True)
    (raw_dir / f"{year:04d}-{month:02d}.xlsx").write_bytes(response.content)

    return response.content


def parse(raw: bytes) -> pd.DataFrame:
    """
    Parse SBI Mutual Fund's all-schemes monthly workbook into the
    standard holdings shape (scheme_name, isin, ticker, quantity, value_inr).

    Parameters
    ----------
    raw : bytes
        The raw .xlsx file content, as returned by fetch().

    Returns
    -------
    pd.DataFrame

    Raises
    ------
    None — rows that don't resolve to a known universe ticker are
    skipped, not raised.
    """
    isin_map = get_isin_to_ticker_map()
    workbook = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)

    records = []
    for sheet_name in workbook.sheetnames:
        if sheet_name == "Index":
            continue
        sheet = workbook[sheet_name]

        scheme_name = None
        for row in sheet.iter_rows(min_row=1, max_row=6, values_only=True):
            if row and len(row) > 3 and row[2] == "SCHEME NAME :":
                scheme_name = row[3]
                break
        if not scheme_name:
            continue

        for row in sheet.iter_rows(min_row=7, values_only=True):
            if len(row) < 7:
                continue
            isin = row[3]
            if not isin or not isinstance(isin, str) or not isin.startswith("IN"):
                continue
            ticker = isin_map.get(isin)
            if not ticker:
                continue
            quantity = row[5] or 0
            market_value_lakhs = row[6] or 0
            records.append(
                {
                    "scheme_name": scheme_name,
                    "isin": isin,
                    "ticker": ticker,
                    "quantity": quantity,
                    "value_inr": market_value_lakhs * 100_000,
                }
            )

    return pd.DataFrame(records, columns=["scheme_name", "isin", "ticker", "quantity", "value_inr"])


# Auto-registers on import — zero cost (no network call, just function
# references); the secondary cross-check is always available without an
# explicit opt-in, same as it was before this module existed standalone.
register_amc(AMC_NAME, fetch, parse)
